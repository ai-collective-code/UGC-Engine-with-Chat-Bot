# -*- coding: utf-8 -*-
"""
Reusable UGC creator outreach engine.

Usage:
    python build_outreach_queue.py <input_xlsx> <config_json> <output_xlsx>

What it does (deterministic, no LLM call needed at runtime):
  1. Loads the two-sheet Apify export (Unique Profiles + All Posts)
  2. Merges each profile with its own posts to pull: phone numbers found
     in captions, and which niche hashtag(s) they post under
  3. Resolves each profile's regional language from location text
  4. Renders a personalized outreach message per profile using the
     client's config (brand name, offer line)
  5. Splits output into an Instagram/Facebook manual-send queue and a
     WhatsApp-ready list (phone number found), with a clickable
     wa.me deep link pre-filled with the message
  6. Writes a formatted Excel workbook with a Status tracking column
  7. Mirrors the run into the local dashboard DB (lib/local_db.py) so
     backend/app.py shows it immediately -- same as uploading through
     the dashboard's Upload Data page.

To onboard a new client: write a new config JSON + point at their
creator export. No code changes needed unless the input sheet shape
or niche vocabulary is genuinely different.

The actual parsing/merge/message-rendering logic lives in
lib/outreach_pipeline.py, shared with the dashboard's upload endpoint --
this script adds the Excel workbook output on top of that.
"""
import sys
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "lib"))
from outreach_pipeline import process_export, sync_to_dashboard_db

STATUS_OPTIONS = ["Not Sent", "Sent", "Replied", "Converted", "Not Interested"]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="2F5233")
BASE_FONT = "Calibri"


def load_config(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def write_sheet(wb, title, df, extra_cols=None):
    ws = wb.create_sheet(title)
    cols = list(df.columns) + (extra_cols or []) + ["Status", "Notes"]
    ws.append(cols)
    for _, row in df.iterrows():
        values = [row.get(c, "") for c in df.columns] + [""] * len(extra_cols or []) + ["Not Sent", ""]
        ws.append(values)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(cols)):
        for cell in row:
            cell.font = Font(name=BASE_FONT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {"Full Name": 18, "Username": 22, "Profile Link": 32, "Location (raw)": 20,
              "Matched State": 16, "Language": 12, "Language Confidence": 10, "Niche": 16,
              "Phone": 14, "Caption Sample": 30, "Personalized Message": 55, "WhatsApp Link": 34,
              "Status": 14, "Notes": 24}
    for i, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 16)

    style_header(ws, len(cols))

    status_col_idx = cols.index("Status") + 1
    status_letter = chr(64 + status_col_idx)
    dv = DataValidation(type="list", formula1=f'"{",".join(STATUS_OPTIONS)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{status_letter}2:{status_letter}{ws.max_row}")

    return ws


def main():
    input_path, config_path, output_path = sys.argv[1], sys.argv[2], sys.argv[3]
    cfg = load_config(config_path)

    ig_df, wa_df = process_export(input_path, cfg)

    try:
        client_key = os.path.splitext(os.path.basename(config_path))[0]
        sync_to_dashboard_db(cfg, client_key, ig_df, wa_df)
    except Exception as e:
        print(f"WARNING: could not sync this run to the dashboard DB: {e}")

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "IG_FB_DM_Queue", ig_df)
    write_sheet(wb, "WhatsApp_Ready", wa_df, extra_cols=["WhatsApp Link"])

    summary = wb.create_sheet("Summary", 0)
    summary["A1"] = f"{cfg['client_name']} — {cfg['campaign_name']}"
    summary["A1"].font = Font(bold=True, size=14)
    stats = [
        ("Total creators", len(ig_df)),
        ("Instagram/Facebook DM queue", len(ig_df)),
        ("WhatsApp-ready (phone found in captions)", len(wa_df)),
        ("Language: Hindi", (ig_df["Language"] == "Hindi").sum()),
        ("Language: Gujarati", (ig_df["Language"] == "Gujarati").sum()),
        ("Language: Marathi", (ig_df["Language"] == "Marathi").sum()),
        ("Language: Telugu", (ig_df["Language"] == "Telugu").sum()),
        ("Language: Odia", (ig_df["Language"] == "Odia").sum()),
        ("Language: Bengali", (ig_df["Language"] == "Bengali").sum()),
        ("Language: Assamese", (ig_df["Language"] == "Assamese").sum()),
        ("Location resolved (high confidence)", (ig_df["Language Confidence"] == "high").sum()),
        ("Location unresolved (default language applied)", (ig_df["Language Confidence"] == "none").sum()),
    ]
    for i, (label, val) in enumerate(stats, start=3):
        summary.cell(row=i, column=1, value=label).font = Font(bold=True, name=BASE_FONT)
        summary.cell(row=i, column=2, value=int(val)).font = Font(name=BASE_FONT)
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 12

    wb.save(output_path)
    print(json.dumps({
        "status": "success",
        "total_creators": len(ig_df),
        "whatsapp_ready": len(wa_df),
        "output": output_path,
    }, indent=2))


if __name__ == "__main__":
    main()
