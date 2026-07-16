"""
Sends WhatsApp template messages to everyone in WhatsApp_Ready with
Status == "Not Sent", marks them Sent on success.

Requires env vars (put in .env, never commit):
  WHATSAPP_ACCESS_TOKEN
  WHATSAPP_PHONE_NUMBER_ID
  WHATSAPP_TEMPLATE_NAME     (must already be approved in Business Manager)
  WHATSAPP_TEMPLATE_LANG     (e.g. "hi" or "en_US" -- must match approval)

Template variable mapping assumes a template like:
  "Namaste {{1}}! Aapke {{2}} posts dekhe... {{3}}"
Adjust `build_components()` below to match YOUR approved template's
actual variable count/order -- this is the one thing that must match
exactly or the send fails.

Usage:
  python send_whatsapp.py <outreach_output.xlsx>
"""
import os
import sys
import requests
from openpyxl import load_workbook

from dotenv import load_dotenv
load_dotenv()

ACCESS_TOKEN = os.environ["WHATSAPP_ACCESS_TOKEN"]
PHONE_NUMBER_ID = os.environ["WHATSAPP_PHONE_NUMBER_ID"]
TEMPLATE_NAME = os.environ["WHATSAPP_TEMPLATE_NAME"]
TEMPLATE_LANG = os.environ.get("WHATSAPP_TEMPLATE_LANG", "en_US")

API_URL = f"https://graph.facebook.com/v20.0/{PHONE_NUMBER_ID}/messages"


def build_components(name, niche):
    """Map creator data to template {{1}}, {{2}}... -- edit to match
    your actual approved template."""
    first_name = (name or "").split()[0] if name else "Bhai"
    return [{
        "type": "body",
        "parameters": [
            {"type": "text", "text": first_name},
            {"type": "text", "text": niche or "construction"},
        ],
    }]


def send_template(phone, name, niche):
    payload = {
        "messaging_product": "whatsapp",
        "to": phone if phone.startswith("91") else f"91{phone}",
        "type": "template",
        "template": {
            "name": TEMPLATE_NAME,
            "language": {"code": TEMPLATE_LANG},
            "components": build_components(name, niche),
        },
    }
    headers = {"Authorization": f"Bearer {ACCESS_TOKEN}", "Content-Type": "application/json"}
    resp = requests.post(API_URL, json=payload, headers=headers, timeout=15)
    return resp.status_code == 200, resp.text


def main():
    path = sys.argv[1]
    wb = load_workbook(path)
    ws = wb["WhatsApp_Ready"]

    headers = {cell.value: cell.column for cell in ws[1]}
    phone_col = headers["Phone"]
    name_col = headers["Full Name"]
    niche_col = headers["Niche"]
    status_col = headers["Status"]
    notes_col = headers["Notes"]

    for row in ws.iter_rows(min_row=2):
        status = row[status_col - 1].value
        if status != "Not Sent":
            continue
        phone = row[phone_col - 1].value
        name = row[name_col - 1].value
        niche = row[niche_col - 1].value
        ok, resp = send_template(str(phone), name, niche)
        row[status_col - 1].value = "Sent" if ok else "Failed"
        row[notes_col - 1].value = "" if ok else resp[:200]
        print(f"{name} ({phone}): {'OK' if ok else 'FAILED - ' + resp[:150]}")

    wb.save(path)
    print("Done. Workbook updated.")


if __name__ == "__main__":
    main()
