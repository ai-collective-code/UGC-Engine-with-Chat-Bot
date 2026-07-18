# -*- coding: utf-8 -*-
"""
Local monitoring dashboard for the UGC Outreach Engine.

Serves the frontend (frontend/) and a small JSON API, both backed by the
same local database (lib/local_db.py) that build_outreach_queue.py and the
negotiator webhooks (whatsapp_ai_agent.py, instagram_webhook.py) write to.
This process doesn't run the negotiator itself -- it's a read/monitor +
light-control surface over data those other processes produce.

Run:
  pip install -r requirements.txt
  python backend/app.py
Then open http://localhost:8000 in a browser.
"""
import os
import re
import sys
import json
import csv
import io
import tempfile
import threading

from dotenv import load_dotenv
load_dotenv()

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "lib"))
import local_db
import outreach_pipeline
import apify_client
import profile_analyzer
import content_qc
import supabase_bridge
import phone_capture
import message_variations
import creator_verification
import language_intelligence

from flask import Flask, jsonify, request, send_from_directory, Response

# Apify actor that backs in-app Instagram scraping (apify/instagram-reel-scraper).
# Override with APIFY_INSTAGRAM_ACTOR if you fork or swap the actor.
INSTAGRAM_ACTOR = os.environ.get("APIFY_INSTAGRAM_ACTOR", "apify~instagram-reel-scraper")
# Apify actor for hashtag-based discovery (apify/instagram-hashtag-scraper): finds
# posts/reels carrying a hashtag, so you can source creators you don't already know.
# Its items expose the same ownerUsername/caption/hashtags fields the reel scraper
# does, so store_apify_items imports them unchanged.
INSTAGRAM_HASHTAG_ACTOR = os.environ.get("APIFY_INSTAGRAM_HASHTAG_ACTOR", "apify~instagram-hashtag-scraper")
_scrape_imported = set()  # run_ids already imported, so re-polling doesn't re-import
_scrape_imported_lock = threading.Lock()  # guards the set under Flask's threaded server

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FRONTEND_DIR = os.path.join(_REPO_ROOT, "frontend")
CONFIG_DIR = os.path.join(_REPO_ROOT, "config")


def _to_int(value, default):
    """int() that can't 500 a request: bad/missing input falls back to default."""
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _safe_cell(v):
    """Excel formula-injection guard for exports: prefix cells that would
    execute as formulas (=, +, -, @) so they open as inert text."""
    s = "" if v is None else str(v)
    return "'" + s if s[:1] in ("=", "+", "-", "@") else s


# Control chars openpyxl refuses to write (IllegalCharacterError) -- scraped
# captions occasionally carry them.
_XLSX_BAD_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

app = Flask(__name__, static_folder=FRONTEND_DIR, static_url_path="")
local_db.init_db()


@app.route("/")
def index():
    return send_from_directory(FRONTEND_DIR, "index.html")


# --- clients ---------------------------------------------------------------

@app.route("/api/clients", methods=["GET"])
def api_list_clients():
    return jsonify(local_db.list_clients())


@app.route("/api/clients/<int:client_id>", methods=["GET"])
def api_get_client(client_id):
    client = local_db.get_client(client_id)
    if not client:
        return jsonify({"error": "not found"}), 404
    return jsonify(client)


@app.route("/api/clients", methods=["POST"])
def api_create_client():
    """Create a new client: writes config/<client_key>.json (so the existing
    CLI scripts + webhooks keep working unmodified via CLIENT_CONFIG) AND a
    DB row, so it shows up in the dashboard immediately."""
    payload = request.get_json(silent=True) or {}
    client_key = (payload.get("client_key") or "").strip().lower().replace(" ", "_")
    if not client_key or not all(c.isalnum() or c == "_" for c in client_key):
        return jsonify({"error": "client_key must be alphanumeric/underscore, e.g. 'acme_corp'"}), 400

    config_path = os.path.join(CONFIG_DIR, f"{client_key}.json")

    client_name = payload.get("client_name") or client_key
    max_voucher = _to_int(payload.get("max_voucher_inr") or 2000, None)
    opening_voucher = _to_int(payload.get("opening_voucher_inr") or max_voucher, None)
    if max_voucher is None or opening_voucher is None:
        return jsonify({"error": "max_voucher_inr and opening_voucher_inr must be numbers"}), 400

    config = {
        "client_name": client_name,
        "campaign_name": payload.get("campaign_name", ""),
        "brand_display_name": payload.get("brand_display_name") or client_name,
        "offer_line": {"value": payload.get("offer_line", "")},
        "default_language": payload.get("default_language", "Hindi"),
        "phone_regex": r"(?:\+?91[\-\s]?)?[6-9]\d{9}",
        "input_sheet": {"profiles_sheet": "Unique Profiles", "posts_sheet": "All Posts"},
        "channels": {
            "instagram_dm": bool(payload.get("instagram_enabled", True)),
            "whatsapp": bool(payload.get("whatsapp_enabled", True)),
            "facebook_dm": bool(payload.get("facebook_enabled", False)),
        },
        "compliance_note": payload.get(
            "compliance_note",
            "Per ASCI influencer guidelines (India), any paid/product-for-content "
            "collaboration should be disclosed by the creator (#ad / #collab) once "
            "content goes live.",
        ),
        "negotiation": {
            "opening_voucher_inr": opening_voucher,
            "max_voucher_inr": max_voucher,
            "voucher_type": payload.get("voucher_type", "Amazon Voucher"),
            "reimbursement": payload.get("reimbursement", ""),
            "deliverables": payload.get("deliverables", ""),
            "human_confirm_before_close": True,
        },
    }

    os.makedirs(CONFIG_DIR, exist_ok=True)
    # "x" mode makes the existence check atomic -- two simultaneous creates
    # can't both write the same config file.
    try:
        with open(config_path, "x", encoding="utf-8") as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
    except FileExistsError:
        return jsonify({"error": f"config/{client_key}.json already exists"}), 409

    client_id = local_db.upsert_client_from_config(config, client_key)
    return jsonify(local_db.get_client(client_id)), 201


# --- creators ----------------------------------------------------------

@app.route("/api/creators", methods=["GET"])
def api_list_creators():
    return jsonify(local_db.list_creators(
        client_id=request.args.get("client_id", type=int),
        status=request.args.get("status"),
        channel=request.args.get("channel"),
        source_platform=request.args.get("source_platform"),
        whatsapp_ready=request.args.get("whatsapp_ready") == "1",
    ))


@app.route("/api/whatsapp-ready", methods=["GET"])
def api_whatsapp_ready():
    """The WhatsApp funnel: every creator with a phone number, whatever
    platform they came from -- the automatic final-contact dashboard."""
    return jsonify(local_db.list_whatsapp_ready(
        client_id=request.args.get("client_id", type=int),
        source_platform=request.args.get("source_platform"),
        status=request.args.get("status"),
    ))


@app.route("/api/platform-breakdown", methods=["GET"])
def api_platform_breakdown():
    return jsonify(local_db.platform_breakdown(client_id=request.args.get("client_id", type=int)))


@app.route("/api/creators/<int:creator_id>", methods=["PATCH"])
def api_update_creator(creator_id):
    payload = request.get_json(silent=True) or {}
    status = payload.get("status")
    ops_stage = payload.get("ops_stage")
    if not status and ops_stage is None:
        return jsonify({"error": "status or ops_stage is required"}), 400
    if status:
        if local_db.update_creator_status(creator_id, status, payload.get("notes")) == 0:
            return jsonify({"error": "creator not found"}), 404
    if ops_stage is not None:
        if local_db.update_creator_ops_stage(creator_id, ops_stage) == 0:
            return jsonify({"error": "creator not found"}), 404
    return jsonify({"status": "ok"})


@app.route("/api/ops-pipeline", methods=["GET"])
def api_ops_pipeline():
    """Post-deal fulfillment board: every Converted creator, tracked through
    product purchase -> content posted -> verified -> payout."""
    return jsonify(local_db.list_ops_pipeline(client_id=request.args.get("client_id", type=int)))


@app.route("/api/ops-pipeline/export", methods=["GET"])
def api_ops_pipeline_export():
    """Finance handoff (V3): a payout-ready CSV of every creator whose content
    has been verified, for the finance/ops team to action manually. No
    payment API is wired up yet (RazorpayX/Cashfree/Xoxoday all require
    business KYC), so this is the handoff point until that's set up."""
    rows = local_db.payout_ready_export(client_id=request.args.get("client_id", type=int))
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Creator", "Username", "Phone", "Niche", "Source", "Language", "Fulfillment Stage", "Deal Note", "Content URL", "QC Score", "QC Verdict"])
    for r in rows:
        writer.writerow([_safe_cell(v) for v in (
            r.get("full_name", ""), r.get("username", ""), r.get("phone", ""), r.get("niche", ""),
            r.get("source_platform", ""), r.get("language", ""), r.get("ops_stage", ""),
            r.get("deal_note", ""), r.get("content_url", ""), r.get("qc_score", ""), r.get("qc_verdict", ""),
        )])
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=payout_ready_export.csv"},
    )


@app.route("/api/language-breakdown", methods=["GET"])
def api_language_breakdown():
    return jsonify(local_db.language_breakdown(client_id=request.args.get("client_id", type=int)))


@app.route("/api/creators/export", methods=["GET"])
def api_creators_export():
    """Download the current creator queue (e.g. a fresh Apify scrape) as an
    .xlsx, formatted like the CLI's build_outreach_queue.py output so it's
    familiar to whoever works the spreadsheet day-to-day."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    rows = local_db.list_creators(
        client_id=request.args.get("client_id", type=int),
        status=request.args.get("status"),
        channel=request.args.get("channel"),
        source_platform=request.args.get("source_platform"),
    )

    columns = [
        ("full_name", "Name"),
        ("username", "Username"),
        ("profile_link", "Profile Link"),
        ("phone", "Phone"),
        ("whatsapp_link", "WhatsApp Link"),
        ("location_raw", "Location"),
        ("language", "Language"),
        ("niche", "Niche"),
        ("source_platform", "Source"),
        ("channel", "Channel"),
        ("status", "Status"),
        ("personalized_message", "Personalized Message"),
        ("caption_sample", "Caption Sample"),
        ("created_at", "Scraped At"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Creators"

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for row_idx, r in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(columns, start=1):
            val = r.get(key, "")
            if isinstance(val, str):
                val = _XLSX_BAD_CHARS.sub("", _safe_cell(val))
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx, (key, label) in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        longest = max([len(label)] + [len(str(r.get(key, "") or "")) for r in rows], default=len(label))
        ws.column_dimensions[letter].width = min(max(longest + 2, 12), 60)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=creators_export.xlsx"},
    )


ALLOWED_UPLOAD_EXTENSIONS = {".xlsx", ".xls"}
VALID_PLATFORMS = {"instagram", "facebook", "youtube", "linkedin"}


def _validate_upload_request(form, files):
    """Shared client/platform/file/config validation for /api/upload and
    /api/upload/preview. Returns (client, cfg, file, ext, error_response).
    error_response is a (jsonify(...), status) tuple, or None if validation passed."""
    client_id = form.get("client_id", type=int)
    if not client_id:
        return None, None, None, None, (jsonify({"error": "client_id is required"}), 400)

    platform = (form.get("platform") or "instagram").strip().lower()
    if platform not in VALID_PLATFORMS:
        return None, None, None, None, (jsonify({"error": f"platform must be one of {sorted(VALID_PLATFORMS)}"}), 400)

    client = local_db.get_client(client_id)
    if not client:
        return None, None, None, None, (jsonify({"error": "unknown client_id"}), 404)

    file = files.get("file")
    if not file or not file.filename:
        return None, None, None, None, (jsonify({"error": "no file uploaded"}), 400)

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        return None, None, None, None, (jsonify({"error": "only .xlsx/.xls files are supported"}), 400)

    config_path = os.path.join(CONFIG_DIR, f"{client['client_key']}.json")
    if not os.path.exists(config_path):
        return None, None, None, None, (jsonify({"error": f"config/{client['client_key']}.json not found on disk"}), 500)
    try:
        with open(config_path, encoding="utf-8") as f:
            cfg = json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        return None, None, None, None, (jsonify({"error": f"config/{client['client_key']}.json is corrupt or unreadable: {e}"}), 500)

    return client, cfg, file, ext, None


@app.route("/api/upload/preview", methods=["POST"])
def api_upload_preview():
    """Dry-run an upload: parse the file and return the detected column
    mapping, header row, row count, phone-match count, and a 5-row sample --
    without writing anything to the DB. Lets the user confirm the mapping
    looks right before committing via /api/upload."""
    client, cfg, file, ext, err = _validate_upload_request(request.form, request.files)
    if err:
        return err

    platform = (request.form.get("platform") or "instagram").strip().lower()
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=ext)
    os.close(tmp_fd)
    try:
        file.save(tmp_path)
        preview = outreach_pipeline.preview_upload(tmp_path, cfg, platform)
    except Exception as e:
        return jsonify({"error": f"failed to preview file: {e}"}), 400
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass  # Windows can hold the handle briefly after pandas reads it; harmless

    return jsonify(preview)


@app.route("/api/upload", methods=["POST"])
def api_upload_creators():
    """Upload a creator export (.xlsx) for a chosen client and source platform
    (instagram/facebook/youtube/linkedin). Runs the parsing/language/message
    pipeline, tags every creator with the platform, and writes them straight
    into the local DB. Any creator with a phone number is automatically routed
    into the WhatsApp dashboard -- no separate step."""
    client, cfg, file, ext, err = _validate_upload_request(request.form, request.files)
    if err:
        return err

    client_id = request.form.get("client_id", type=int)
    platform = (request.form.get("platform") or "instagram").strip().lower()

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=ext)
    os.close(tmp_fd)  # release the handle before Werkzeug re-opens it to save (Windows can't share it)
    try:
        file.save(tmp_path)
        result = outreach_pipeline.store_upload(cfg, client_id, tmp_path, platform)
    except Exception as e:
        return jsonify({"error": f"failed to process file: {e}"}), 400
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass  # Windows can hold the handle briefly after pandas reads it; harmless

    return jsonify({
        "status": "success",
        "platform": platform,
        "total_creators": result["total"],
        "whatsapp_ready": result["whatsapp_ready"],
    })


# --- Apify live scraping -----------------------------------------------------

def _load_client_cfg(client):
    """Load a client's config JSON off disk (for the scrape/import pipeline).
    Raises RuntimeError on a corrupt/unreadable file so callers surface a clean
    JSON error instead of an HTML 500."""
    config_path = os.path.join(CONFIG_DIR, f"{client['client_key']}.json")
    if not os.path.exists(config_path):
        return None
    try:
        with open(config_path, encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        raise RuntimeError(f"config/{client['client_key']}.json is corrupt or unreadable: {e}")


@app.route("/api/scrape", methods=["POST"])
def api_scrape_start():
    """Kick off an Apify Instagram scrape. Returns a run_id to poll. Costs Apify
    credits, so it's gated behind an explicit button + a per-profile results cap."""
    if not apify_client.is_configured():
        return jsonify({"error": "APIFY_TOKEN is not set in .env — add it, then restart the dashboard."}), 400

    payload = request.get_json(silent=True) or {}
    client_id = payload.get("client_id")
    if not client_id or not local_db.get_client(client_id):
        return jsonify({"error": "valid client_id is required"}), 400

    platform = (payload.get("platform") or "instagram").strip().lower()
    if platform != "instagram":
        return jsonify({"error": "In-app scraping is wired for Instagram only right now."}), 400

    targets = [str(t).strip() for t in (payload.get("targets") or []) if str(t).strip()]
    if not targets:
        return jsonify({"error": "Enter at least one username, profile URL, or reel URL."}), 400

    results_limit = max(1, min(_to_int(payload.get("results_limit"), 25), 200))
    run_input = {
        "username": targets,
        "resultsLimit": results_limit,
        "skipPinnedPosts": False,
        "skipTrialReels": False,
        "includeSharesCount": False,
        "includeTranscript": False,
        "includeDownloadedVideo": False
    }

    try:
        run = apify_client.start_run(INSTAGRAM_ACTOR, run_input)
    except Exception as e:
        return jsonify({"error": str(e)}), 502
    return jsonify({"run_id": run["run_id"], "status": run["status"]})


@app.route("/api/scrape/hashtag", methods=["POST"])
def api_scrape_hashtag_start():
    """Kick off an Apify Instagram HASHTAG scrape (apify/instagram-hashtag-scraper).
    Discovers posts/reels carrying the given hashtags and returns a run_id to poll
    via the same /api/scrape/status endpoint. Costs Apify credits (~$2.60 / 1,000
    results), so it's gated behind an explicit button + a per-hashtag results cap."""
    if not apify_client.is_configured():
        return jsonify({"error": "APIFY_TOKEN is not set in .env — add it, then restart the dashboard."}), 400

    payload = request.get_json(silent=True) or {}
    client_id = payload.get("client_id")
    if not client_id or not local_db.get_client(client_id):
        return jsonify({"error": "valid client_id is required"}), 400

    # Accept "#tiling", "tiling", full hashtag URLs — normalize to bare tags.
    raw = payload.get("hashtags") or []
    hashtags = []
    for h in raw:
        tag = str(h).strip().lstrip("#").rstrip("/")
        if "/" in tag:  # e.g. instagram.com/explore/tags/tiling
            tag = tag.rsplit("/", 1)[-1]
        if tag:
            hashtags.append(tag)
    if not hashtags:
        return jsonify({"error": "Enter at least one hashtag."}), 400

    # The hashtag actor calls this resultsType; "posts" covers posts + reels.
    results_type = (payload.get("content_type") or "posts").strip().lower()
    if results_type not in {"posts", "reels", "stories"}:
        results_type = "posts"

    results_limit = max(1, min(_to_int(payload.get("results_limit"), 20), 200))
    run_input = {
        "hashtags": hashtags,
        "resultsType": results_type,
        "resultsLimit": results_limit,
    }

    try:
        run = apify_client.start_run(INSTAGRAM_HASHTAG_ACTOR, run_input)
    except Exception as e:
        return jsonify({"error": str(e)}), 502
    return jsonify({"run_id": run["run_id"], "status": run["status"]})


@app.route("/api/scrape/status", methods=["GET"])
def api_scrape_status():
    """Poll a scrape run. On SUCCEEDED, fetch the dataset, import the creators
    (once), and return the counts. Import is idempotent (upsert), and guarded
    so re-polling doesn't re-import."""
    run_id = request.args.get("run_id")
    client_id = request.args.get("client_id", type=int)
    platform = (request.args.get("platform") or "instagram").strip().lower()
    if not run_id or not client_id:
        return jsonify({"error": "run_id and client_id are required"}), 400

    try:
        run = apify_client.get_run(run_id)
    except Exception as e:
        return jsonify({"error": str(e)}), 502

    status = run["status"]
    if status in apify_client.TERMINAL_BAD:
        return jsonify({"status": status, "error": f"Scrape run {status.lower()}"})
    if status not in apify_client.TERMINAL_OK:
        return jsonify({"status": status})  # READY / RUNNING — keep polling

    # SUCCEEDED
    with _scrape_imported_lock:
        if run_id in _scrape_imported:
            return jsonify({"status": status, "imported": True})
        # Claim the run before the slow fetch/import so a concurrent poll
        # can't import it twice; discarded below if the import fails.
        _scrape_imported.add(run_id)

    client = local_db.get_client(client_id)
    try:
        cfg = _load_client_cfg(client) if client else None
        if not cfg:
            with _scrape_imported_lock:
                _scrape_imported.discard(run_id)
            return jsonify({"status": status, "error": "client config not found on disk"}), 500
        items = apify_client.fetch_items(run["dataset_id"])
        result = outreach_pipeline.store_apify_items(cfg, client_id, items, platform)
    except Exception as e:
        with _scrape_imported_lock:
            _scrape_imported.discard(run_id)
        return jsonify({"status": status, "error": f"import failed: {e}"}), 502

    return jsonify({
        "status": status,
        "imported": True,
        "scraped_items": len(items),
        "total_creators": result["total"],
        "whatsapp_ready": result["whatsapp_ready"],
    })


# --- message variations (spam-safe outreach rewrites) ----------------------

@app.route("/api/message-variations", methods=["POST"])
def api_message_variations():
    """Rewrite one outreach message into N same-meaning versions so the first
    (manual) DM to each creator isn't identical text -- Meta flags duplicates."""
    payload = request.get_json(silent=True) or {}
    sentence = str(payload.get("sentence") or "").strip()
    count = payload.get("count", 15)
    if not sentence:
        return jsonify({"error": "sentence is required"}), 400
    try:
        variations = message_variations.generate_variations(sentence, count=count)
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 502
    except (ValueError, TypeError) as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"variations": variations})


# --- profile analyzer ------------------------------------------------------

@app.route("/api/analyze-profile", methods=["POST"])
def api_analyze_profile():
    payload = request.get_json(silent=True) or {}
    client_id = payload.get("client_id")
    username = str(payload.get("username") or "").strip().lower()
    
    if not username or not client_id:
        return jsonify({"error": "client_id and username are required"}), 400
        
    cached = local_db.get_profile_analysis(username, client_id)
    if cached:
        return jsonify({"status": "CACHED", "result": json.loads(cached["analysis_data"])})
        
    run_input = {
        "username": [username],
        "resultsLimit": 5,
        "skipPinnedPosts": False,
        "skipTrialReels": False,
        "includeSharesCount": False,
        "includeTranscript": False,
        "includeDownloadedVideo": False
    }
    
    try:
        run = apify_client.start_run(INSTAGRAM_ACTOR, run_input)
    except Exception as e:
        return jsonify({"error": str(e)}), 502
    return jsonify({"run_id": run["run_id"], "status": run["status"]})


@app.route("/api/analyze-profile/status", methods=["GET"])
def api_analyze_profile_status():
    run_id = request.args.get("run_id")
    client_id = request.args.get("client_id", type=int)
    username = request.args.get("username")
    
    if not run_id or not client_id or not username:
        return jsonify({"error": "run_id, client_id, and username are required"}), 400
        
    try:
        run = apify_client.get_run(run_id)
    except Exception as e:
        return jsonify({"error": str(e)}), 502

    status = run["status"]
    if status in apify_client.TERMINAL_BAD:
        return jsonify({"status": status, "error": f"Scrape run {status.lower()}"})
    if status not in apify_client.TERMINAL_OK:
        return jsonify({"status": status})

    try:
        items = apify_client.fetch_items(run["dataset_id"])
        if not items:
            return jsonify({"status": status, "error": "Profile not found or private."})
            
        client = local_db.get_client(client_id)
        cfg = _load_client_cfg(client) if client else {}
        
        profile_data = items[0].copy()
        profile_data["latestPosts"] = items
        
        analysis = profile_analyzer.analyze_profile_data(profile_data, cfg)
        
        local_db.save_profile_analysis(
            username=username, 
            profile_data=json.dumps(profile_data), 
            analysis_data=json.dumps(analysis), 
            client_id=client_id
        )
        
        return jsonify({"status": status, "result": analysis})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"status": status, "error": f"Analysis failed: {str(e)}"}), 502


# --- creator verification (roadmap points 46-50) ---------------------------

@app.route("/api/verify-creator", methods=["POST"])
def api_verify_creator():
    """Kick off an Apify scrape of a creator, then (on status poll) run the
    fraud/trust checks. Mirrors the analyze-profile flow: POST returns a run_id
    to poll via /api/verify-creator/status. Reputation history (point 50) is
    merged in at status time from the local DB -- no scrape needed for that."""
    payload = request.get_json(silent=True) or {}
    client_id = payload.get("client_id")
    username = str(payload.get("username") or "").strip().lstrip("@").lower()
    if not username or not client_id:
        return jsonify({"error": "client_id and username are required"}), 400
    if not apify_client.is_configured():
        return jsonify({"error": "APIFY_TOKEN is not set in .env — add it, then restart the dashboard."}), 400

    run_input = {
        "username": [username],
        "resultsLimit": 6,
        "skipPinnedPosts": False,
        "skipTrialReels": False,
        "includeSharesCount": False,
        "includeTranscript": False,
        "includeDownloadedVideo": False,
    }
    try:
        run = apify_client.start_run(INSTAGRAM_ACTOR, run_input)
    except Exception as e:
        return jsonify({"error": str(e)}), 502
    return jsonify({"run_id": run["run_id"], "status": run["status"]})


@app.route("/api/verify-creator/status", methods=["GET"])
def api_verify_creator_status():
    run_id = request.args.get("run_id")
    client_id = request.args.get("client_id", type=int)
    username = request.args.get("username")
    if not run_id or not client_id or not username:
        return jsonify({"error": "run_id, client_id, and username are required"}), 400

    try:
        run = apify_client.get_run(run_id)
    except Exception as e:
        return jsonify({"error": str(e)}), 502

    status = run["status"]
    if status in apify_client.TERMINAL_BAD:
        return jsonify({"status": status, "error": f"Scrape run {status.lower()}"})
    if status not in apify_client.TERMINAL_OK:
        return jsonify({"status": status})

    try:
        items = apify_client.fetch_items(run["dataset_id"])
        if not items:
            return jsonify({"status": status, "error": "Profile not found or private."})

        client = local_db.get_client(client_id)
        cfg = _load_client_cfg(client) if client else {}

        profile_data = items[0].copy()
        profile_data["latestPosts"] = items

        report = creator_verification.verify_creator(profile_data, cfg)
        # Point 50: fold in any stored reputation history for this creator.
        report = creator_verification.merge_reputation(
            report, local_db.list_reputation(username)
        )
        return jsonify({"status": status, "result": report})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"status": status, "error": f"Verification failed: {str(e)}"}), 502


# --- creator reputation (roadmap point 50) ---------------------------------

@app.route("/api/reputation", methods=["GET"])
def api_get_reputation():
    username = str(request.args.get("username") or "").strip()
    if not username:
        return jsonify({"error": "username is required"}), 400
    return jsonify(local_db.list_reputation(username))


@app.route("/api/reputation", methods=["POST"])
def api_add_reputation():
    payload = request.get_json(silent=True) or {}
    username = str(payload.get("username") or "").strip()
    event = str(payload.get("event") or "").strip()
    detail = str(payload.get("detail") or "").strip()
    if not username or not event:
        return jsonify({"error": "username and event are required"}), 400
    try:
        local_db.add_reputation(username, event, detail, client_id=payload.get("client_id"))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"status": "ok", "records": local_db.list_reputation(username)})


# --- language intelligence (roadmap points 42-45) --------------------------

@app.route("/api/language/supported", methods=["GET"])
def api_language_supported():
    return jsonify({"languages": language_intelligence.SUPPORTED_LANGUAGES})


@app.route("/api/language/detect-style", methods=["POST"])
def api_language_detect_style():
    """Point 42/43: read a creator's message, report language/register/dialect."""
    payload = request.get_json(silent=True) or {}
    text = str(payload.get("text") or "").strip()
    if not text:
        return jsonify({"error": "text is required"}), 400
    try:
        return jsonify(language_intelligence.detect_style(text))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/language/translate", methods=["POST"])
def api_language_translate():
    """Point 44: translate text into a target language, romanized by default."""
    payload = request.get_json(silent=True) or {}
    text = str(payload.get("text") or "").strip()
    target = str(payload.get("target_language") or "").strip()
    romanized = payload.get("romanized", True)
    if not text or not target:
        return jsonify({"error": "text and target_language are required"}), 400
    try:
        return jsonify(language_intelligence.translate(text, target, romanized=bool(romanized)))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/language/check-sensitivity", methods=["POST"])
def api_language_check_sensitivity():
    """Point 45: flag culturally/regionally inappropriate content before send."""
    payload = request.get_json(silent=True) or {}
    text = str(payload.get("text") or "").strip()
    if not text:
        return jsonify({"error": "text is required"}), 400
    try:
        return jsonify(language_intelligence.check_sensitivity(
            text,
            target_language=payload.get("target_language"),
            region=payload.get("region"),
        ))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 502


# --- dealers (V2 PIN-code lookup) --------------------------------------------

@app.route("/api/dealers/upload", methods=["POST"])
def api_dealers_upload():
    """Upload a dealer/stockist list (.xlsx) for a client -- powers the
    nearest-dealer lookup once real data is available."""
    client_id = request.form.get("client_id", type=int)
    if not client_id:
        return jsonify({"error": "client_id is required"}), 400
    if not local_db.get_client(client_id):
        return jsonify({"error": "unknown client_id"}), 404

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "no file uploaded"}), 400
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        return jsonify({"error": "only .xlsx/.xls files are supported"}), 400

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=ext)
    os.close(tmp_fd)
    try:
        file.save(tmp_path)
        result = outreach_pipeline.store_dealer_upload(client_id, tmp_path)
    except Exception as e:
        return jsonify({"error": f"failed to process file: {e}"}), 400
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    return jsonify({"status": "success", "total_dealers": result["total"]})


@app.route("/api/dealers", methods=["GET"])
def api_list_dealers():
    return jsonify(local_db.list_dealers(client_id=request.args.get("client_id", type=int)))


@app.route("/api/dealers/nearest", methods=["GET"])
def api_nearest_dealer():
    client_id = request.args.get("client_id", type=int)
    if not client_id:
        return jsonify({"error": "client_id is required"}), 400
    dealer = local_db.find_nearest_dealer(
        client_id,
        pincode=request.args.get("pincode"),
        state=request.args.get("state"),
    )
    if not dealer:
        return jsonify({"error": "no matching dealer found — upload a dealer list first, or none match this pincode/state"}), 404
    return jsonify(dealer)


# --- content QC (V2) --------------------------------------------------------

@app.route("/api/qc/run", methods=["POST"])
def api_qc_run():
    """Kick off an AI QC pass on a creator's submitted content URL (a reel or
    post link). Scrapes the content via Apify, then scores it with Gemini
    against the campaign's deliverable brief + QC rubric."""
    payload = request.get_json(silent=True) or {}
    client_id = payload.get("client_id")
    creator_id = payload.get("creator_id")
    content_url = str(payload.get("content_url") or "").strip()
    if not (client_id and creator_id and content_url):
        return jsonify({"error": "client_id, creator_id, and content_url are required"}), 400
    if not apify_client.is_configured():
        return jsonify({"error": "APIFY_TOKEN is not set in .env — add it, then restart the dashboard."}), 400

    run_input = {
        "username": [content_url],
        "resultsLimit": 1,
        "skipPinnedPosts": False,
        "skipTrialReels": False,
        "includeSharesCount": False,
        "includeTranscript": False,
        "includeDownloadedVideo": False,
    }
    try:
        run = apify_client.start_run(INSTAGRAM_ACTOR, run_input)
    except Exception as e:
        return jsonify({"error": str(e)}), 502
    return jsonify({"run_id": run["run_id"], "status": run["status"]})


@app.route("/api/qc/status", methods=["GET"])
def api_qc_status():
    run_id = request.args.get("run_id")
    client_id = request.args.get("client_id", type=int)
    creator_id = request.args.get("creator_id", type=int)
    content_url = request.args.get("content_url")
    if not (run_id and client_id and creator_id and content_url):
        return jsonify({"error": "run_id, client_id, creator_id, and content_url are required"}), 400

    try:
        run = apify_client.get_run(run_id)
    except Exception as e:
        return jsonify({"error": str(e)}), 502

    status = run["status"]
    if status in apify_client.TERMINAL_BAD:
        return jsonify({"status": status, "error": f"Scrape run {status.lower()}"})
    if status not in apify_client.TERMINAL_OK:
        return jsonify({"status": status})

    try:
        items = apify_client.fetch_items(run["dataset_id"])
        if not items:
            return jsonify({"status": status, "error": "Content not found, private, or removed."})

        client = local_db.get_client(client_id)
        cfg = _load_client_cfg(client) if client else {}
        result = content_qc.analyze_content(items[0], cfg)
        if local_db.save_content_qc(creator_id, content_url, result) == 0:
            return jsonify({"status": status, "error": "creator not found"}), 404
        return jsonify({"status": status, "result": result})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"status": status, "error": f"QC analysis failed: {str(e)}"}), 502


# --- conversations -----------------------------------------------------

# Channels a creator can share a WhatsApp number in and get auto-routed to the
# WhatsApp funnel. WhatsApp itself is the destination, so it's not a source.
_CAPTURE_CHANNELS = {"instagram", "facebook"}


def _route_number_to_whatsapp(client_id, contact_id, number):
    """Upsert a creator who shared a WhatsApp number in an IG/Messenger chat
    into the WhatsApp funnel, attached to the given (active) client. A creator
    with a phone shows up in the WhatsApp section automatically -- this is the
    final outreach stage. Idempotent via upsert_whatsapp_contact's (client_id,
    username, channel) upsert, so re-scanning the same thread won't duplicate
    the row -- and, unlike a full insert_creators upsert, it only touches the
    phone/whatsapp_link columns, so an enriched imported row isn't wiped.
    Returns the number on success, or None if it couldn't attribute/store."""
    if not (client_id and number and contact_id):
        return None
    if not local_db.get_client(client_id):
        return None
    local_db.upsert_whatsapp_contact(
        client_id, contact_id, contact_id, number,
        outreach_pipeline.wa_link(number, ""),
        profile_link=f"https://instagram.com/{contact_id}",
    )
    return number


def _default_capture_client_id(active_client_id=None):
    """Which client an auto-captured WhatsApp number attaches to. Prefer the
    actively-selected client; otherwise fall back to the first client so numbers
    are captured even under "All clients" (IG DMs aren't tied to a client)."""
    if active_client_id:
        return active_client_id
    clients = local_db.list_clients()
    return clients[0]["id"] if clients else None


HIDDEN_CONVS_SETTING = "hidden_conversations"


def _get_hidden_conversations():
    """Set of "channel::contact_id" keys the operator has hidden from the
    conversations list. Persisted in the settings table so the hide applies on
    every browser. View-only -- messages are never deleted."""
    try:
        return set(json.loads(local_db.get_setting(HIDDEN_CONVS_SETTING, "[]")))
    except Exception:
        return set()


def _save_hidden_conversations(keys):
    local_db.set_setting(HIDDEN_CONVS_SETTING, json.dumps(sorted(keys)))


def _conv_key(channel, contact_id):
    return f"{channel}::{contact_id}"


@app.route("/api/conversations", methods=["GET"])
def api_list_conversations():
    channel = request.args.get("channel")
    active_client_id = request.args.get("client_id", type=int)
    convos = local_db.list_conversations(
        client_id=active_client_id,
        channel=channel,
    )
    # Merge in Instagram DMs from the chatbot's Supabase (read-only bridge).
    # They're client-agnostic, so they always appear regardless of the client
    # filter; only the channel filter can hide them. Sorted newest-first with
    # the local rows by last activity.
    if channel in (None, supabase_bridge.CHANNEL):
        ig = supabase_bridge.list_instagram_conversations()
        # Auto-route: any creator who dropped a WhatsApp number in a DM gets
        # upserted into the WhatsApp funnel automatically. IG DMs aren't tied to
        # a client, so with "All clients" we attribute to the default (first)
        # client rather than skipping — the number is captured either way.
        capture_client_id = _default_capture_client_id(active_client_id)
        for c in ig:
            num = c.get("detected_whatsapp")
            if num and capture_client_id:
                if _route_number_to_whatsapp(capture_client_id, c["contact_id"], num):
                    c["whatsapp_captured"] = True
        convos = convos + ig
        convos.sort(key=lambda c: c.get("last_message_at") or "", reverse=True)
    # Flag operator-hidden conversations (view-only) rather than dropping them,
    # so the frontend can show a "N hidden / Show all" control. No deletion.
    hidden = _get_hidden_conversations()
    for c in convos:
        c["hidden"] = _conv_key(c.get("channel"), c.get("contact_id")) in hidden
    return jsonify(convos)


@app.route("/api/conversations/hide", methods=["POST"])
def api_hide_conversation():
    """Hide a conversation from the list (view-only, reversible). Does NOT
    delete the conversation or any of its messages."""
    payload = request.get_json(silent=True) or {}
    channel = str(payload.get("channel") or "").strip()
    contact_id = str(payload.get("contact_id") or "").strip()
    if not channel or not contact_id:
        return jsonify({"error": "channel and contact_id are required"}), 400
    hidden = _get_hidden_conversations()
    hidden.add(_conv_key(channel, contact_id))
    _save_hidden_conversations(hidden)
    return jsonify({"status": "hidden", "hidden_count": len(hidden)})


@app.route("/api/conversations/unhide", methods=["POST"])
def api_unhide_conversation():
    """Un-hide one conversation, or all of them with {"all": true}."""
    payload = request.get_json(silent=True) or {}
    if payload.get("all"):
        _save_hidden_conversations(set())
        return jsonify({"status": "all_unhidden", "hidden_count": 0})
    channel = str(payload.get("channel") or "").strip()
    contact_id = str(payload.get("contact_id") or "").strip()
    if not channel or not contact_id:
        return jsonify({"error": "channel and contact_id are required (or pass all=true)"}), 400
    hidden = _get_hidden_conversations()
    hidden.discard(_conv_key(channel, contact_id))
    _save_hidden_conversations(hidden)
    return jsonify({"status": "unhidden", "hidden_count": len(hidden)})


@app.route("/api/conversations/messages", methods=["GET"])
def api_conversation_messages():
    channel = request.args.get("channel")
    contact_id = request.args.get("contact_id")
    # capture_client_id: the active client to attribute a captured number to
    # (the sentinel client_id used by Instagram threads can't own a creator).
    capture_client_id = request.args.get("capture_client_id", type=int)
    # Instagram threads live in Supabase, keyed only by contact_id (username).
    if channel == supabase_bridge.CHANNEL:
        if not contact_id:
            return jsonify({"error": "contact_id is required"}), 400
        history = supabase_bridge.get_instagram_history(contact_id)
        # Full-history scan on open catches a number shared in an older thread
        # that fell outside the list view's recent-message batch. Falls back to
        # the default client so capture works even under "All clients".
        capture_client_id = _default_capture_client_id(capture_client_id)
        if capture_client_id:
            num = phone_capture.capture_from_history(history)
            if num:
                _route_number_to_whatsapp(capture_client_id, contact_id, num)
        return jsonify(history)

    client_id = request.args.get("client_id", type=int)
    if not (client_id and channel and contact_id):
        return jsonify({"error": "client_id, channel, and contact_id are required"}), 400
    history = local_db.get_history(client_id, channel, contact_id)
    if channel in _CAPTURE_CHANNELS:
        num = phone_capture.capture_from_history(history)
        if num:
            _route_number_to_whatsapp(client_id, contact_id, num)
    return jsonify(history)


@app.route("/api/conversations/instagram/send", methods=["POST"])
def api_instagram_send():
    """Send a human reply to an Instagram DM (proxied to the chatbot)."""
    data = request.get_json(silent=True) or {}
    contact_id = data.get("contact_id")
    message = (data.get("message") or "").strip()
    if not contact_id or not message:
        return jsonify({"error": "contact_id and message are required"}), 400
    ok, err = supabase_bridge.send_instagram_reply(contact_id, message)
    if not ok:
        return jsonify({"error": err or "Send failed"}), 502
    return jsonify({"status": "sent"})


# --- human actions -------------------------------------------------------

@app.route("/api/human-actions", methods=["GET"])
def api_list_human_actions():
    resolved_param = request.args.get("resolved")
    resolved = None if resolved_param is None else resolved_param == "1"
    return jsonify(local_db.list_human_actions(resolved=resolved))


@app.route("/api/human-actions/<int:action_id>/resolve", methods=["POST"])
def api_resolve_human_action(action_id):
    local_db.resolve_human_action(action_id)
    return jsonify({"status": "ok"})


# --- negotiator control (dashboard kill switch) --------------------------

@app.route("/api/negotiator/status", methods=["GET"])
def api_negotiator_status():
    val = local_db.get_setting("negotiator_paused", "false")
    return jsonify({"paused": val.strip().lower() in ("1", "true", "yes")})


@app.route("/api/negotiator/pause", methods=["POST"])
def api_negotiator_pause():
    local_db.set_setting("negotiator_paused", "true")
    return jsonify({"paused": True})


@app.route("/api/negotiator/resume", methods=["POST"])
def api_negotiator_resume():
    local_db.set_setting("negotiator_paused", "false")
    return jsonify({"paused": False})


# --- activity feed ----------------------------------------------------------

@app.route("/api/activity-feed", methods=["GET"])
def api_activity_feed():
    return jsonify(local_db.activity_feed(
        client_id=request.args.get("client_id", type=int),
        limit=_to_int(request.args.get("limit"), 30),
    ))


@app.route("/api/pipeline-stages", methods=["GET"])
def api_pipeline_stages():
    return jsonify(local_db.pipeline_stages(
        client_id=request.args.get("client_id", type=int),
    ))


# --- stats -----------------------------------------------------------------

@app.route("/api/stats", methods=["GET"])
def api_stats():
    return jsonify(local_db.get_stats(client_id=request.args.get("client_id", type=int)))


if __name__ == "__main__":
    # Bind 0.0.0.0 and honour $PORT so this runs under a host like Render (which
    # injects PORT and routes to it) as well as locally. Production deploys
    # should front this with gunicorn (see requirements.txt) rather than Flask's
    # dev server; this block is the local-dev / fallback entrypoint.
    port = int(os.environ.get("PORT", 8000))
    print("Dashboard DB: PostgreSQL (schema 'ugc')")
    print(f"Starting dashboard on http://0.0.0.0:{port}")
    app.run(host="0.0.0.0", port=port, debug=False)
